# ----------------------------------------------
# File name: RMT_Processing.py
# Author: Jean Mattekatt
# Date: 7/11/2023

# Description: This code reads DDR5 data and converts it into a histogram, tables, boxplot, 
# bit margin (scatterplot), and average bit margin graphs. Prints 'done' in terminal once finished

# Assumption:
#   There is no csv files with the name: vendorName + "_" + marginType + ".csv"
#       vendorName ex: 3rd to last word in folder name, separated by '_'
#                       Everett_EMR_SK_64GB_1DPC --> SK
#                       Everett_EMR_Samsung_64GB_1DPC --> Samsung
#                       Everett_EMR_Micron_64GB_1DPC --> Micron
#       marginType ex: CPU0_RankMargin, CPU0_LaneMargin, CPU1_RankMargin, CPU1_LaneMargin
#   All folders have the same number of files
# ----------------------------------------------

# pip install matplotlib
# pip install scikit-learn
# pip install openpyxl


import os
import numpy as np
import matplotlib.pyplot as plt
import statistics as stats
import csv
from sklearn.utils import resample
import openpyxl


def processData(folders, vendorNames, bootstrap, includeLine, histogram, vendorTable, boxPlot, varTable, bitMarg, comparator):
    """
    processData puts file data into variables and prints done after plotting and saving the graphs

    Args:
        folders (list): contains all folder paths
        vendorNames (list): contains all vendor names
        bootstrap (str): 'Y' if yes, 'N' if no
        includeLine (str): 'Y' if yes, 'N' if no
        histogram (boolean): True if graphs histogram
        vendorTable (boolean): True if creates vendor table
        boxPlot (boolean): True if graphs box plot
        varTable (boolean): True if creates variable table
        bitMarg (boolean): True if graphs bit margin (scatterplot)
        comparator (boolean): True if graphs average bit margin (line graph)

    """ 
    allRankMarginCPU0, allLaneMarginCPU0 = [None] * len(folders), [None] * len(folders)
    allRankMarginCPU1, allLaneMarginCPU1, variableList = [None] * len(folders), [None] * len(folders), [None] * len(folders)
    for i in range(0, len(folders)):
        allRankMarginCPU0[i], allLaneMarginCPU0[i], allRankMarginCPU1[i], allLaneMarginCPU1[i], variableList = readData(os.listdir(folders[i]), folders[i], vendorNames[i])
    
    makeGraphs(allRankMarginCPU0, variableList, vendorNames, includeLine, bootstrap, "CPU0 Rank Margin", histogram, vendorTable, boxPlot, varTable, bitMarg, comparator, len(os.listdir(folders[0])))
    makeGraphs(allLaneMarginCPU0, variableList, vendorNames, includeLine, bootstrap, "CPU0 Lane Margin", histogram, vendorTable, boxPlot, varTable, bitMarg, comparator, len(os.listdir(folders[0])))
    
    makeGraphs(allRankMarginCPU1, variableList, vendorNames, includeLine, bootstrap, "CPU1 Rank Margin", histogram, vendorTable, boxPlot, varTable, bitMarg, comparator, len(os.listdir(folders[0])))
    makeGraphs(allLaneMarginCPU1, variableList, vendorNames, includeLine, bootstrap, "CPU1 Lane Margin", histogram, vendorTable, boxPlot, varTable, bitMarg, comparator, len(os.listdir(folders[0])))

    print("Done")

    
def readData(folder, folderPath, vendorName):
    """
    readData reads files, organizes, and saves data to variables

    Args:
        folder (list): contains all files in folder
        folderPath (str): path of folder
        vendorName (str): name of vendor
        
    Returns:
        list: contains all CPU0 rank margin data
        list: contains all CPU0 lane margin data
        list: contains all CPU1 rank margin data
        list: contains all CPU1 lane margin data
        list: contains all variables

    """ 
    allRankMarginCPU0 = []
    allLaneMarginCPU0 = []

    allRankMarginCPU1 = []
    allLaneMarginCPU1 = []

    variableListCPU0 = ""

    # reads through all files in folder
    for fileName in folder:
        file = open(folderPath + "\\" + fileName, "r")
        text = file.readlines()
        dataCPU0 = ""
        dataCPU1 = ""
        recordCPU0Data = False
        recordCPU1Data = False

        # saves data between 'START_RMT' and 'STOP_RMT'
        for line in text:
            if line.find("START_RMT_N0") != -1:
                recordCPU0Data = True
            if line.find("STOP_RMT_N0") != -1:
                recordCPU0Data = False
            if recordCPU0Data:
                dataCPU0 = dataCPU0 + line
            if line.find("START_RMT_N1") != -1:
                recordCPU1Data = True
            if line.find("STOP_RMT_N1") != -1:
                recordCPU1Data = False
            if recordCPU1Data:
                dataCPU1 = dataCPU1 + line
        file.close()

        # separate each CPU data in lists by margin type 
        rankMarginCPU0, laneMarginCPU0, variableListCPU0 = separateCPU(dataCPU0, str(0), vendorName)
        rankMarginCPU1, laneMarginCPU1, variableListCPU1 = separateCPU(dataCPU1, str(1), vendorName)

        allRankMarginCPU0.extend(rankMarginCPU0)
        allLaneMarginCPU0.extend(laneMarginCPU0)

        allRankMarginCPU1.extend(rankMarginCPU1)
        allLaneMarginCPU1.extend(laneMarginCPU1)

    return allRankMarginCPU0, allLaneMarginCPU0, allRankMarginCPU1, allLaneMarginCPU1, variableListCPU0


def separateCPU(data, cpuNum, vendorName):
    """
    separateCPU separates the rank and lane margin data for each CPU

    Args:
        data (str): all data from folder for one CPU
        cpuNum (str): CPU number
        vendorName (str): name of vendor
        
    Returns:
        list: contains all rank margin data for CPU
        list: contains all lane margin data for CPU

    """ 
    rankMargin = data[data.find("Rank Margin"):data.find("Lane Margin")]
    laneMargin = data[data.find("Lane Margin"):data.find("CA Lane Margin")]

    variableList = rankMargin[rankMargin.find("RxDqs-"):rankMargin.find("\nN" + cpuNum)]
    variableList = list(filter(None, variableList.split(" ")))

    # separate each CPU data in lists by line
    rankMargin = separateMarginData(rankMargin, "CPU" + cpuNum + "_RankMargin", vendorName)
    laneMargin = separateMarginData(laneMargin, "CPU" + cpuNum + "_LaneMargin", vendorName)

    return rankMargin, laneMargin, variableList


def separateMarginData(data, marginType, vendorName):
    """
    separateMarginData separates line of data for each margin

    Args:
        data (str): all data from folder for one type of margin (e.g. rank, lane)
        marginType (str): type of margin
        vendorName (str): name of vendor
        
    Returns:
        list: contains all lines of data for margin type

    """ 
    data = data[data.find("N" + marginType[3]):data.rfind("\nIoLevel")]
    data = data.splitlines()
    for i in range(0, len(data)):
        data[i] = list(filter(None, data[i].split(" ")))
    createCSVFile(vendorName + "_" + marginType + '.csv', data)
    return data


def makeGraphs(allMarginList, variableList, vendorNames, includeLine, bootstrap, marginType, histogram, vendorTable, boxPlot, varTable, bitMarg, comparator, numFiles):    
    """
    makeGraphs plots each type of graph/table that was selected in the GUI

    Args:
        allMarginList (list): contains all data for one margin
        variableList (list): contains all variables
        vendorNames (list): contains all vendor names
        includeLine (str): 'Y' if yes, 'N' if no
        bootstrap (str): 'Y' if yes, 'N' if no
        marginType (str): type of margin
        histogram (boolean): True if graphs histogram
        vendorTable (boolean): True if creates vendor table
        boxPlot (boolean): True if graphs box plot
        varTable (boolean): True if creates variable table
        bitMarg (boolean): True if graphs bit margin (scatterplot)
        comparator (boolean): True if graphs average bit margin (line graph)
        numFiles (int): number of files in folder

    """ 
    if boxPlot:
        boxFig, boxAxs = plt.subplots(2, 4)

    if varTable:
        tableFig, tableAxs = plt.subplots(2, 4)

    if comparator:
        compGraphs = makeCompGraphs(allMarginList[0])

    allMean = []
    allMedian = []
    allSD = []
    allIQR = []
    allMeanSD1 = []
    allMeanSD2 = []
    allMeanSD3 = []

    if boxPlot or varTable or vendorTable:
        for i in range(1, 9):
            columns, x, y, mean, median, sd, iqr, meanSD1, meanSD2, meanSD3 = calculateStats(allMarginList, i, bootstrap)

            if boxPlot:
                makeBoxPlot(columns, vendorNames, variableList[i - 1], boxAxs[x, y], includeLine)

            if varTable:
                makeVarTable(mean, median, sd, iqr, meanSD1, meanSD2, meanSD3, vendorNames, variableList[i-1], tableAxs[x, y])

            if vendorTable:
                allMean.append(mean)
                allMedian.append(median)
                allSD.append(sd)
                allIQR.append(iqr)
                allMeanSD1.append(meanSD1)
                allMeanSD2.append(meanSD2)
                allMeanSD3.append(meanSD3)

        if boxPlot:
            boxFig.subplots_adjust(top=0.85, bottom=0.05, wspace=0.3, hspace=0.3)
            boxFig.suptitle(marginType)
            boxFig.savefig(marginType.replace(" ", "_") + "BoxPlot.pdf")

        if varTable:
            # different spacing between tables according to number of tables
            if len(vendorNames) == 1:
                tableFig.subplots_adjust(top=0.85, bottom=0.05, wspace=1.5, hspace=0.3)
            elif len(vendorNames) == 2:
                tableFig.subplots_adjust(top=0.85, bottom=0.05, wspace=0.8, hspace=0.3)
            elif len(vendorNames) == 3:
                tableFig.subplots_adjust(top=0.85, bottom=0.05, wspace=0.4, hspace=0.3)
            tableFig.suptitle(marginType)
            tableFig.savefig(marginType.replace(" ", "_") + "VarTable.pdf", bbox_inches='tight')

        if vendorTable:
            makeTable(variableList, vendorNames, allMean, allMedian, allSD, allIQR, allMeanSD1, allMeanSD2, allMeanSD3, marginType)


    if histogram or bitMarg or comparator:
        for i in range(0, len(vendorNames)):
            if histogram:
                histFig, histAxs = plt.subplots(2, 4)
            
                for j in range(1, 9):
                    makeHistogram(allMarginList[i], variableList[j - 1], j, histAxs, includeLine, bootstrap)
            
                histFig.subplots_adjust(top=0.9, bottom=0.18, wspace=0.3, hspace=0.75)
                histFig.suptitle(vendorNames[i] + " " + marginType)
                histFig.savefig(vendorNames[i] + "_" + marginType.replace(" ", "_") + "Histogram.pdf")

            if allMarginList[0][0][0].find('L') != -1 and bitMarg:
                makeBitMargin(allMarginList[i], variableList, includeLine, vendorNames[i])

            if allMarginList[0][0][0].find('L') != -1 and comparator:
                makeComparator(allMarginList[i], compGraphs, variableList, includeLine, vendorNames[i], numFiles)


def createCSVFile(fileName, marginData):
    """
    createCSVFile creates a CSV file for one margin

    Args:
        fileName (str): name of csv file
        rankMarginData (list): contains all data for one margin

    """
    if os.path.isfile(fileName):
        with open(fileName, 'a', newline='') as file:
            writer = csv.writer(file)
            writer.writerows(marginData)
    else:
        with open(fileName, 'w', newline='') as file:
            writer = csv.writer(file)
            writer.writerows(marginData)


def makeVarTable(mean, median, sd, iqr, meanSD1, meanSD2, meanSD3, vendorNames, variable, axs):
    """
    makeVarTable creates a table of statistics comparing all vendors for each variable

    Args:
        mean (list): contains mean of each vendor for variable
        median (list): contains median of each vendor for variable
        sd (list): contains standard deviation of each vendor for variable
        iqr (list): contains interquartile range of each vendor for variable
        meanSD1 (list): contains value one standard deviation from mean of each vendor for variable
        meanSD2 (list): contains value two standard deviation from mean of each vendor for variable
        meanSD3 (list): contains value three standard deviation from mean of each vendor for variable
        vendorNames (list): contains all vendor names
        variable (str): name of variable
        axs (axes): graph axes and position of subplot

    """
    axs.axis('tight')
    axs.axis('off')

    row = ["Mean", "Median", "SD", "IQR", "Mean-1SD", "Mean-2SD", "Mean-3SD"]
    axs.table(cellText=[mean, median, sd, iqr, meanSD1, meanSD2, meanSD3], rowLabels=row, colLabels=vendorNames, loc='center')
    axs.set_title(variable)


def makeComparator(marginList, allGraphs, variableList, includeLine, vendor, numFiles):
    """
    makeComparator organizes data and creates an excel containing average bit margin graphs
        
    Args:
        marginList (list): contains all data for one margin
        allGraphs (list): contains axes and figures of all average bit margin graphs
        variableList (list): contains all variables
        includeLine (str): 'Y' if yes, 'N' if no
        vendor (str): name of vendor
        numFiles (int): number of files in folder

    """
    avgData = {}

    # organize and input data into dictionary
    for i in range(0, len(marginList)):
        tabName = marginList[i][0][:marginList[i][0].find('C')+2]
        if not tabName in avgData:
            avgData[tabName] = {}

        graphTitle = marginList[i][0][:marginList[i][0].rfind('.')]
        if not graphTitle in avgData[tabName]:
            avgData[tabName][graphTitle] = [],[],[],[],[],[],[],[]
        
        laneNum = int(marginList[i][0][marginList[i][0].rfind('.') + 2:-1])
        for k in range(0, 8):
            if len(avgData[tabName][graphTitle][k]) < 40:
                avgData[tabName][graphTitle][k].append(abs(int(marginList[i][k+1])))
            else:
                avgData[tabName][graphTitle][k][laneNum] += abs(int(marginList[i][k+1]))

    # find average of data
    for tabName in avgData:
        for graphTitle in avgData[tabName]:
            for i in range(0, 8):
                avgData[tabName][graphTitle] = list(avgData[tabName][graphTitle])
                avgData[tabName][graphTitle][i] = [round(num/numFiles,2) for num in avgData[tabName][graphTitle][i]]

    # create excel file
    wb = openpyxl.Workbook()
    del wb['Sheet']
    
    for tabName in avgData:
        ws = wb.create_sheet(tabName)
        graphPos = 1

        for graphTitle in avgData[tabName]:
            for graphNum in range(0, 8):
                x = 0
                y = graphNum
                if graphNum > 3:
                    x = 1
                    y = graphNum % 4

                # create each graph
                allGraphs[tabName][graphTitle][0][x, y].plot(range(0, 40), avgData[tabName][graphTitle][graphNum], label=vendor)
                allGraphs[tabName][graphTitle][0][x, y].grid(linestyle='dotted')
                allGraphs[tabName][graphTitle][0][x, y].set_title(variableList[graphNum])
                allGraphs[tabName][graphTitle][0][x, y].legend(fontsize='5', loc='upper right')

                if includeLine == "Y":
                    allGraphs[tabName][graphTitle][0][x, y].axhline(y=6, linestyle='dashed')

            allGraphs[tabName][graphTitle][1].suptitle(graphTitle)
            allGraphs[tabName][graphTitle][1].set_figwidth(25)
            allGraphs[tabName][graphTitle][1].subplots_adjust(top=0.85, bottom=0.1, wspace=0.3, hspace=0.3)
            allGraphs[tabName][graphTitle][1].savefig(vendor+'_'+graphTitle+"Comparator.png")
            img = openpyxl.drawing.image.Image(vendor+'_'+graphTitle+'Comparator.png')

            # add graph to excel
            ws.add_image(img, anchor='A'+str(graphPos))
            
            graphPos = graphPos + 26

    wb.save('LaneComparison.xlsx')
    wb.close()


def makeCompGraphs(marginList):
    """
    makeCompGraphs creates axes and figures of all average bit margin graphs
        
    Args:
        marginList (list): contains all data for one margin

    Returns:
        list: contains axes and figures of all average bit margin graphs

    """
    allCompGraphs = {}

    # create all graphs for average bit margin comparison
    for i in range(0, len(marginList)):
        tabName = marginList[i][0][:marginList[i][0].find('C')+2]
        if not tabName in allCompGraphs:
            allCompGraphs[tabName] = {}

        graphTitle = marginList[i][0][:marginList[i][0].rfind('.')]
        if not graphTitle in allCompGraphs[tabName]:
            compFig, compAxs = plt.subplots(2, 4)
            allCompGraphs[tabName][graphTitle] = [compAxs, compFig]

    return allCompGraphs


def makeBitMargin(marginList, variableList, includeLine, vendor):
    """
    makeBitMargin organizes data and creates an excel containing bit margin graphs
        
    Args:
        marginList (list): contains all data for one margin
        variableList (list): contains all variables
        includeLine (str): 'Y' if yes, 'N' if no
        vendor (str): name of vendor

    """
    allData = {}
    allGraphs = {}

    # organize and input data into dictionary, create all bit margin graphs
    for i in range(0, len(marginList)):
        tabName = marginList[i][0][:marginList[i][0].find('C')+2]
        if not tabName in allData:
            allData[tabName] = {}
            allGraphs[tabName] = {}

        graphTitle = marginList[i][0][:marginList[i][0].rfind('.')]
        if not graphTitle in allData[tabName]:
            allData[tabName][graphTitle] = [],[],[],[],[],[],[],[],[]
            
            bitMargFig, bitMargAxs = plt.subplots(2, 4)
            allGraphs[tabName][graphTitle] = [bitMargAxs, bitMargFig]

        laneNum = int(marginList[i][0][marginList[i][0].rfind('.') + 2:-1])
        for k in range(0, 9):
            if k == 0:
                allData[tabName][graphTitle][k].append(laneNum)
            else:
                allData[tabName][graphTitle][k].append(abs(int(marginList[i][k])))

    # create excel file
    wb = openpyxl.Workbook()
    del wb['Sheet']

    for tabName in allData:
        ws = wb.create_sheet(tabName)
        graphPos = 1

        for graphTitle in allData[tabName]:
            for graphNum in range(1, 9):
                x = 0
                y = graphNum - 1
                if graphNum > 4:
                    x = 1
                    y = (graphNum - 1) % 4

                # create each graph
                allGraphs[tabName][graphTitle][0][x, y].scatter(allData[tabName][graphTitle][0], allData[tabName][graphTitle][graphNum])
                allGraphs[tabName][graphTitle][0][x, y].grid(linestyle='dotted')
                allGraphs[tabName][graphTitle][0][x, y].set_title(variableList[graphNum-1])

                if includeLine == "Y":
                    allGraphs[tabName][graphTitle][0][x, y].axhline(y=6, linestyle='dashed')

            allGraphs[tabName][graphTitle][1].suptitle(graphTitle)
            allGraphs[tabName][graphTitle][1].set_figwidth(25)
            allGraphs[tabName][graphTitle][1].subplots_adjust(top=0.85, bottom=0.1, wspace=0.3, hspace=0.3)
            allGraphs[tabName][graphTitle][1].savefig(vendor+'_'+graphTitle+".png")
            img = openpyxl.drawing.image.Image(vendor+'_'+graphTitle+'.png')

            # add graph to excel
            ws.add_image(img, anchor='A'+str(graphPos))
                
            graphPos = graphPos + 26

    wb.save(vendor+'BitMargin.xlsx')
    wb.close()


def makeBoxPlot(columns, vendorNames, variable, axs, includeLine):
    """
    makeBoxPlot creates a box plot comparing all vendors for each variable
        
    Args:
        columns (list): contains all data for variable
        vendorNames (list): contains all vendor names
        variable (str): variable name
        axs (axes): graph axes and position of subplot
        includeLine (str): 'Y' if yes, 'N' if no

    """
    axs.boxplot(columns, labels=vendorNames)

    axs.tick_params(axis='both', which='major', labelsize=5)
    axs.set_title(variable)
    axs.grid(linestyle='dotted')

    if includeLine == "Y":
        axs.axhline(y=6, linestyle='dotted')
    axs.plot()


def makeTable(variableList, vendorNames, allMean, allMedian, allSD, allIQR, allMeanSD1, allMeanSD2, allMeanSD3, marginType):
    """
    makeTable creates a table of statistics comparing variables for each vendor
        
    Args:
        variableList (list): contains all variables
        vendorNames (list): contains all vendor names
        allMean (list): contains all mean of data
        allMedian (list): contains all median of data
        allSD (list): contains all standard deviation of data
        allIQR (list): contains all interquartile range of data
        allMeanSD1 (list): contains value one standard of deviaton from each mean of data
        allMeanSD2 (list): contains value one standard of deviaton from each mean of data
        allMeanSD3 (list): contains value one standard of deviaton from each mean of data
        marginType (str): type of margin

    """
    dataNum = len(allMean[0])
    rowTot, colTot = dataNum, 1

    if dataNum > 2:
        rowTot, colTot = 2, 2
    tableFig, tableAx = plt.subplots(rowTot, colTot, squeeze=False)
    for col in range(0, colTot):
        for row in range(0, rowTot):
            tableAx[row, col].axis('tight')
            tableAx[row, col].axis('off')

    rows = ["Mean", "Median", "SD", "IQR", "Mean-1SD", "Mean-2SD", "Mean-3SD"]
    for tableNum in range(0, dataNum):
        col = 0
        row = tableNum
        if tableNum + 1 > 2:
            col = 1
            row = tableNum % 2

        meanRow, medianRow, sdRow, iqrRow, meanSD1Row, meanSD2Row, meanSD3Row = [], [], [], [], [], [], []
        for colNum in range(0, len(allMean)):
            meanRow.append(allMean[colNum][tableNum])
            medianRow.append(allMedian[colNum][tableNum])
            sdRow.append(allSD[colNum][tableNum])
            iqrRow.append(allIQR[colNum][tableNum])
            meanSD1Row.append(allMeanSD1[colNum][tableNum])
            meanSD2Row.append(allMeanSD2[colNum][tableNum])
            meanSD3Row.append(allMeanSD3[colNum][tableNum])
        tableAx[row, col].table(cellText=[meanRow, medianRow, sdRow, iqrRow, meanSD1Row, meanSD2Row, meanSD3Row],
                                rowLabels=rows, colLabels=variableList, loc='center')
        tableAx[row, col].set_title(vendorNames[tableNum])

    tableFig.suptitle(marginType)
    tableFig.subplots_adjust(top=0.85, bottom=0.04, hspace=0.75)
    tableFig.savefig(marginType.replace(" ", "") + "VendorTable.pdf", bbox_inches='tight')


def calculateStats(marginList, graphNum, bootstrap):
    """
    calculateStats calculates the statistics of each vendor for each variable
        
    Args:
        marginList (list): contains all data for one margin
        graphNum (int): position of subplot
        bootstrap (str): 'Y' if yes, 'N' if no

        columns, x, y, mean, median, stdev, iqr, meanSD1, meanSD2, meanSD3

    Returns:
        list: contains all data for variable
        int: x position of subplots
        int: y position of subplots
        list: contains mean of each vendor for variable
        list: contains median of each vendor for variable
        list: contains standard deviation of each vendor for variable
        list: contains interquartile range of each vendor for variable
        list: contains value one standard deviation from mean of each vendor for variable
        list: contains value two standard deviation from mean of each vendor for variable
        list: contains value three standard deviation from mean of each vendor for variable

    """   
    x = 0
    y = graphNum - 1
    if graphNum > 4:
        x = 1
        y = (graphNum - 1) % 4
    
    columns, bins = [], []
    mean, median, stdev, iqr, meanSD1, meanSD2, meanSD3 = [], [], [], [], [], [], []

    for i in range(0, len(marginList)):

        columns.append([])
        for j in range(0, len(marginList[i])):
            columns[i].append(abs(int(marginList[i][j][graphNum])))

        if bootstrap == "Y":
            columns[i] = resample(columns[i], replace=True, n_samples=1000, random_state=1)

        bins.append(range(min(columns[i]), max(columns[i]) + 2))

        mean.append(round(stats.mean(columns[i]), 4))
        median.append(round(stats.median(columns[i]), 4))
        stdev.append(round(stats.stdev(columns[i]), 4))
        iqr.append(round(np.subtract(*np.percentile(columns[i], [75, 25])), 4))
        meanSD1.append(round(mean[i] - stdev[i], 4))
        meanSD2.append(round(mean[i] - (2 * stdev[i]), 4))
        meanSD3.append(round(mean[i] - (3 * stdev[i]), 4))

    return columns, x, y, mean, median, stdev, iqr, meanSD1, meanSD2, meanSD3


def makeHistogram(marginList, variable, graphNum, axs, includeLine, bootstrap): 
    """
    makeHistogram creates a histogram of each variable for each vendor
        
    Args:
        marginList (list): contains all data for one margin
        variable (str): variable name
        graphNum (int): position of subplot
        axs (axes): graph axes
        includeLine (str): 'Y' if yes, 'N' if no
        bootstrap (str): 'Y' if yes, 'N' if no

    """   
    x = 0
    y = graphNum - 1
    if graphNum > 4:
        x = 1
        y = (graphNum - 1) % 4

    columns = []

    for i in range(0, len(marginList)):
        columns.append(abs(int(marginList[i][graphNum])))

    if bootstrap == "Y":
        columns = resample(columns, replace=True, n_samples=1000, random_state=1)

    bins = range(min(columns), max(columns) + 2)

    axs[x, y].hist(columns, bins, edgecolor='black')

    mean = round(stats.mean(columns), 4)
    stdev = round(stats.stdev(columns), 4)

    axs[x, y].tick_params(axis='both', which='major', labelsize=5)
    axs[x, y].grid(linestyle='dotted')
    axs[x, y].set_title(variable)

    if includeLine == "Y":
        axs[x, y].axvline(x=6, linestyle='dotted')
    axs[x, y].plot()

    col_labels = ['mean', 'standard\ndeviation']
    table_vals = [mean, stdev]
    axs[x, y].table(cellText=[table_vals], colLabels=col_labels, bbox=[0, -0.5, 1, 0.35])
